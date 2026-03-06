from __future__ import annotations
from fastapi import APIRouter, Request, Form, Depends
from fastapi.responses import RedirectResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime, timedelta, date as dt_date
from ..core.session import get_session
from ..core.config import settings
from ..services.accounting import add_payment, list_payments, delete_payment, month_settlement, update_order_discount, PAY_METHODS
from ..services.pricing import FISH_TYPES, FISH_SIZES, ensure_defaults_for_date, get_catalog, set_catalog, get_prices, set_prices
from ..services.customers import list_customers, create_customer, rotate_token, delete_customer
from ..services.orders import list_orders, get_order
from ..services.dispatch import list_drivers, list_vehicles

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

def require_owner(request: Request):
    s = get_session(request)
    if not s or s.get("role") != "owner":
        return None
    return s


# Backward/typo compatibility: earlier draft used `require_admin`.
def require_admin(request: Request):
    return require_owner(request)

@router.get("/admin")
def admin_home(request: Request):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    today = datetime.now().date().isoformat()
    tomorrow = (datetime.now().date() + timedelta(days=1)).isoformat()
    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request, "session": s,
        "today": today, "tomorrow": tomorrow,
        "orders_today": list_orders(today),
        "orders_tomorrow": list_orders(tomorrow),
    })

@router.get("/admin/customers")
def admin_customers(request: Request):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    base_url = (settings.BASE_URL or str(request.base_url).rstrip("/")).rstrip("/")
    return templates.TemplateResponse("admin_customers.html", {
        "request": request, "session": s,
        "customers": list_customers(),
        "base_url": base_url
    })

@router.post("/admin/customers/add")
def admin_customer_add(request: Request, name: str = Form(...), region: str = Form(""), phone: str = Form(""), address: str = Form("")):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    create_customer(name.strip(), region.strip(), phone.strip(), address.strip())
    return RedirectResponse("/admin/customers", status_code=303)

@router.post("/admin/customers/{customer_id}/rotate")
def admin_customer_rotate(request: Request, customer_id: int):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    rotate_token(customer_id)
    return RedirectResponse("/admin/customers", status_code=303)

@router.post("/admin/customers/{customer_id}/delete")
def admin_customer_delete(request: Request, customer_id: int):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    delete_customer(customer_id)
    return RedirectResponse("/admin/customers", status_code=303)

@router.get("/admin/dispatch")
def admin_dispatch(request: Request, date: str | None = None):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)

    base = datetime.now().date()
    today = base.isoformat()
    yesterday = (base - timedelta(days=1)).isoformat()
    tomorrow = (base + timedelta(days=1)).isoformat()

    if not date:
        date = today

    return templates.TemplateResponse("admin_dispatch.html", {
        "request": request, "session": s, "date": date,
        "today": today, "yesterday": yesterday, "tomorrow": tomorrow,
        "orders": list_orders(date),
        "drivers": list_drivers(),
        "vehicles": list_vehicles(True)
    })

@router.get("/admin/orders/{order_id}")
def admin_order_detail(request: Request, order_id: int):
    s = require_owner(request)
    if not s:
        return RedirectResponse("/admin/login", status_code=303)
    return templates.TemplateResponse("admin_order_detail.html", {"request": request, "session": s, "order": get_order(order_id)})


@router.get("/admin/catalog", response_class=HTMLResponse)
def admin_catalog(request: Request, date: str | None = None, user=Depends(require_admin)):
    d = date or dt_date.today().isoformat()
    ensure_defaults_for_date(d)
    catalog = get_catalog(d)
    return templates.TemplateResponse("admin_catalog.html", {"request": request, "date": d, "fish_types": FISH_TYPES, "fish_sizes": FISH_SIZES, "catalog": catalog})

@router.post("/admin/catalog", response_class=HTMLResponse)
async def admin_catalog_save(request: Request, date: str = Form(...), user=Depends(require_admin)):
    """Save per-day catalog toggles.

    NOTE: FastAPI does not support capturing arbitrary form fields via **kwargs.
    If we keep **form in the function signature, FastAPI treats it as a required
    *query* param named "form" and returns 422.
    """
    form = await request.form()
    enabled_map = {}
    for ft in FISH_TYPES:
        for sz in FISH_SIZES:
            enabled_map[(ft, sz)] = (f"en__{ft}__{sz}" in form)
    set_catalog(date, enabled_map)
    return RedirectResponse(url=f"/admin/catalog?date={date}", status_code=303)

@router.get("/admin/prices", response_class=HTMLResponse)
def admin_prices(request: Request, date: str | None = None, user=Depends(require_admin)):
    d = date or dt_date.today().isoformat()
    ensure_defaults_for_date(d)
    catalog = get_catalog(d)
    prices = get_prices(d)
    return templates.TemplateResponse("admin_prices.html", {"request": request, "date": d, "fish_types": FISH_TYPES, "fish_sizes": FISH_SIZES, "catalog": catalog, "prices": prices})

@router.post("/admin/prices", response_class=HTMLResponse)
async def admin_prices_save(request: Request, date: str = Form(...), user=Depends(require_admin)):
    """Save per-day unit prices.

    Same reason as admin_catalog_save(): read form fields from request.form().
    """
    form = await request.form()
    price_map = {}
    for ft in FISH_TYPES:
        for sz in FISH_SIZES:
            key = f"p__{ft}__{sz}"
            raw = (form.get(key) or "").strip().replace(",", "")
            if raw == "":
                price_map[(ft, sz)] = None
            else:
                try:
                    price_map[(ft, sz)] = int(raw)
                except Exception:
                    price_map[(ft, sz)] = None
    set_prices(date, price_map)
    return RedirectResponse(url=f"/admin/prices?date={date}", status_code=303)


@router.get("/admin/settlement", response_class=HTMLResponse)
def admin_settlement(request: Request, month: str | None = None, user=Depends(require_admin)):
    m = month or dt_date.today().strftime("%Y-%m")
    data = month_settlement(m)
    return templates.TemplateResponse("admin_settlement.html", {"request": request, "month": m, "data": data})

@router.get("/admin/settlement.xlsx")
def admin_settlement_excel(month: str | None = None, user=Depends(require_admin)):
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    # Lazy import so the app still boots even if openpyxl is missing.
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HTMLResponse("openpyxl 패키지가 필요합니다.", status_code=500)

    m = month or dt_date.today().strftime("%Y-%m")
    data = month_settlement(m)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "월간_요약"
    totals = data["totals"]
    ws1.append(["월", data["month"]])
    ws1.append(["총 주문건수", totals["order_count"]])
    ws1.append(["총 물량(kg)", totals["total_kg"]])
    ws1.append(["총 매출(합계)", totals["gross_sum"]])
    ws1.append(["총 할인", totals["discount_sum"]])
    ws1.append(["총 순매출", totals["net_sum"]])
    ws1.append(["총 입금", totals["paid_sum"]])
    ws1.append(["가격 미정 주문건", totals["unknown_price_count"]])

    ws2 = wb.create_sheet("거래처별")
    headers = [
        "거래처",
        "지역",
        "주문건수",
        "총물량(kg)",
        "매출합계",
        "할인합계",
        "순매출",
        "입금합계",
        "가격미정건",
        "현재미수금",
    ]
    ws2.append(headers)

    for row in data["rows"]:
        ws2.append([
            row["customer_name"],
            row["customer_region"],
            row["order_count"],
            row["total_kg"],
            row["gross_sum"],
            row["discount_sum"],
            row["net_sum"],
            row["paid_sum"],
            row["unknown_price_count"],
            row["balance"],
        ])

    for ws in (ws1, ws2):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"settlement_{m}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
@router.post("/admin/orders/{order_id}/discount", response_class=HTMLResponse)
def admin_order_discount_save(request: Request, order_id: int, discount_amount: int = Form(0), user=Depends(require_admin)):
    if discount_amount < 0:
        discount_amount = 0
    update_order_discount(int(order_id), int(discount_amount))
    return RedirectResponse(url=f"/admin/orders/{order_id}", status_code=303)



